import datetime
import sys, os
import time
from dataclasses import dataclass, field
import json, time
import openpyxl
from thefuzz import fuzz
import pandas as pd

row_tag = [['转出日期', '日期'],'时间1','银行','转出账号','转出金额',['转入', '转入日期'],['转入时间', '时间', '时间2'],'付款账号','收到账号','转入金额','备注']
bill_id = 0
@dataclass(repr=False)
class personage:
    name: str = field(default=None)
    account: list = field(default=None)
    def __init__(self , name, account=None):
        self.name = name

@dataclass(repr=False)
class bill_item:
    id: int = field(default=None)
    date: datetime = field(default=None)
    from_ac: str = field(default=None)
    payer_info: personage = field(default=None)
    payee_info: personage = field(default=None)
    type: int = field(default=None)
    figure: float = field(default=None)
    def __init__(self, date=None, payer_info=None, from_ac=None, payee_info=None, type=None, figure=None):
        global bill_id
        self.id = bill_id
        bill_id +=1
        self.date = date
        self.from_ac = from_ac
        self.payer_info = payer_info
        self.payee_info = payee_info
        self.type = type
        self.figure = figure

@dataclass(repr=False)
class personage_info(personage):
    bill: list[bill_item] = field(default=None)
    total_count: int = field(default=0)
    # name_another: str = field(default=None)

    def __init__(self, _name):
        super().__init__(_name)
        self.bill = []
    # def __post_init__(self):
    #     self.total_price = self.price * self.count
    #
    # def take(self):
    #     self.total_count = self.total_count - self.count
    # def addcount(self, _count):
    #     self.total_count = self.total_count + 1 * _count

def softID(_it, _itl, _score):
    res_score = {}
    for it in _itl:
        ratio_it = fuzz.partial_ratio(it, _it)
        res_score[it] = ratio_it

    max_key_value = max(res_score.items(), key=lambda item: item[1])
    # print(sorted_dict)
    return max_key_value[0], max_key_value[1]

def create_folder_if_not_exists(folder_path):
    # 判断文件夹是否存在
    if not os.path.exists(folder_path):
        # 递归创建文件夹
        os.makedirs(folder_path)
        print(f"文件夹 '{folder_path}' 创建成功")
    else:
        print(f"文件夹 '{folder_path}' 已存在")

def generate_timestamp_filename(extension):
    # 获取当前时间
    now = datetime.datetime.now()
    # 格式化时间字符串
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    # 生成文件名
    filename = f"{timestamp}.{extension}"
    return filename

def loadxlsx(_sheet, type, ac_info, all_bill):
    if _sheet.max_row < 2:
        return None

    tag_map = {}
    tags = list(_sheet[2])
    index = 0
    for it in row_tag:
        for cell in tags[index:]:  # 获取第 2 行的所有单元格
            if cell.value and cell.value in it:
                if isinstance(it, list):
                    tag_map[it[-1]] = cell.column
                else:
                    tag_map[it] = cell.column
                index += 1
                break
    print(tag_map)
    for i in range(_sheet.max_row + 1):
        if i < 3:
            continue
        date1 = _sheet.cell(row=i, column=tag_map['日期']).value if '日期' in tag_map.keys() else None
        time1 = _sheet.cell(row=i, column=tag_map['时间1']).value if '时间1' in tag_map.keys() else None
        send1 = _sheet.cell(row=i, column=tag_map['银行']).value if '银行' in tag_map.keys() else None
        recv1 = _sheet.cell(row=i, column=tag_map['转出账号']).value if '转出账号' in tag_map.keys() else None
        p1 = _sheet.cell(row=i, column=tag_map['转出金额']).value if '转出金额' in tag_map.keys() else None
        date2 = _sheet.cell(row=i, column=tag_map['转入日期']).value if '转入日期' in tag_map.keys() else None
        time2 = _sheet.cell(row=i, column=tag_map['时间2']).value if '时间2' in tag_map.keys() else None
        send2 = _sheet.cell(row=i, column=tag_map['付款账号']).value if '付款账号' in tag_map.keys() else None
        recv2 = _sheet.cell(row=i, column=tag_map['收到账号']).value if '收到账号' in tag_map.keys() else None
        p2 = _sheet.cell(row=i, column=tag_map['转入金额']).value if '转入金额' in tag_map.keys() else None

        if p1 is None and p2 is None:
            continue
        try:
            p1=int(p1) if p1 is not None else None
            p2=int(p2) if p2 is not None else None
        except:
            print('{}用户的金额 的单元格 异常 : L {}'.format(ac_info.name, i))
        if not isinstance(p1, (int, float)) and not isinstance(p2, (int, float)):
            print('{}用户的金额 的单元格 异常 : L {}'.format(ac_info.name, i))
            continue

        if not isinstance(date1, datetime.datetime) and date1:
            if '.' in date1:
                date1 = datetime.datetime.strptime(date1, '%Y.%m.%d')
            elif '-' in date1:
                date1 = datetime.datetime.strptime(date1, '%Y-%m-%d')
            else:
                date1 = datetime.datetime.strptime(date1.replace("\t", "").replace('号','日'), '%Y年%m月%d日')
        if not isinstance(date2, datetime.datetime) and date2:
            if '.' in date2:
                date2 = datetime.datetime.strptime(date2, '%Y.%m.%d')
            elif '-' in date2:
                date2 = datetime.datetime.strptime(date2, '%Y-%m-%d')
            else:
                date2 = datetime.datetime.strptime(date2.replace("\t", "").replace('号','日'), '%Y年%m月%d日')

        if not isinstance(time1, datetime.time) and time1:
            if ':' in time1:
                time1 = datetime.datetime.strptime(time1, '%H:%M:%S')
            else:
                time1 = datetime.time(0,0)
        if not isinstance(time2, datetime.time) and time2:
            if ':' in time2:
                time2 = datetime.datetime.strptime(time2, '%H:%M:%S')
            else:
                time2 = datetime.time(0,0)

        if isinstance(p1, int):
            time1 = datetime.time(0,0) if time1 is not isinstance(time1, datetime.datetime) else time1
            date1 = datetime.datetime(0,0, 0) if date1 is None else date1
            date1 = date1.replace(hour=time1.hour, minute=time1.minute, second=time1.second)
            bill_A = bill_item(payer_info=personage(send1), payee_info=personage(recv1), from_ac=ac_info.name,
                               figure=p1, date=date1, type=0)
            ac_info.bill.append(bill_A)
            all_bill.append(bill_A)

        if isinstance(p2, int) :
            time1 = datetime.time(0,0) if time1 is not isinstance(time1, datetime.datetime) else time1
            time2 = time1 if time2 is None else time2
            date2 = date1 if date2 is None else date2
            date2 = date2.replace(hour=time2.hour, minute=time2.minute, second=time2.second)
            bill_B = bill_item(payer_info=personage(send2), payee_info=personage(recv2), from_ac=ac_info.name,
                               figure=p2, date=date2, type=1)
            ac_info.bill.append(bill_B)
            all_bill.append(bill_B)

def get_path_formR(_data_p):
    for root, dirs, _ in os.walk(_data_p):
        if len(dirs) < 1:
            print('目录未找到人员文件夹')
            return None, None
        for ac_dir in dirs:
            for ac_p, _, files in os.walk(os.path.join(root, ac_dir)):
                xlsxDs = []
                for it_f in files:
                    if it_f.endswith('xlsx'):
                        xlsxDs.append(it_f)
                if len(xlsxDs) < 1:
                    print('{} 人员文件夹 没有xlsx 数据'.format(ac_p))
                for it_f in xlsxDs:
                    t_file = os.path.join(ac_p, it_f)
                    print('load file : {}'.format(t_file))
                    yield ac_dir, t_file


def load_data(_data_p):
    Collated_D = {}
    A_Bill_D = []
    for ac_name, file in get_path_formR(_data_p):
        if ac_name is None:
            continue
        if ac_name not in Collated_D.keys():
            print('create ac : {}'.format(ac_name))
            Collated_D[ac_name] = personage_info(ac_name)

        # if ac_name == '行长':
        #     continue
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook['data']
        loadxlsx(sheet, 0, Collated_D[ac_name], A_Bill_D)

    A_Bill_D = sorted(A_Bill_D, key=lambda x: x.date)
    print('load data done')
    return Collated_D, A_Bill_D

def export_data(_data, _o_p):
    Collated_D, A_Bill_D = _data

    create_folder_if_not_exists(_o_p)
    t_o_p = os.path.join(_o_p, generate_timestamp_filename('xlsx'))
    print('输出文件为：{}'.format(t_o_p))

    wb = openpyxl.Workbook()
    # 选择默认的工作表
    # wsheet = wb.active
    ws_1 = wb.create_sheet('gyq')

    Date_B_D = {}
    for it in A_Bill_D:
        if it.date.date() not in Date_B_D.keys():
            Date_B_D[it.date.date()] = []
        Date_B_D[it.date.date()].append(it)

    res_date = []
    res_time = []
    res_bz = []
    res_price = []
    res_ac_tg = []
    total_price = [0, 0]

    calculate_date = []
    calculate_time = []
    calculate_price = []
    calculate_bz = []
    calculate_ac_tg = []
    for it in Date_B_D.keys():
        T_Day = [0, 0]
        for itb in Date_B_D[it]:
            res_date.append(itb.date.date())
            res_time.append(itb.date.time())
            res_price.append(itb.figure)
            res_bz.append('send:{},recv:{},{}'.format(itb.payer_info.name, itb.payee_info.name, itb.type))
            if itb.type:
                res_ac_tg.append('{}->{}'.format('耿', itb.from_ac))
                T_Day[-1] += itb.figure
                total_price[-1] += itb.figure
            else:
                res_ac_tg.append('{}->{}'.format(itb.from_ac, '耿'))
                T_Day[0] += itb.figure
                total_price[0] += itb.figure

        calculate_date.append(it)
        calculate_time.append(datetime.time(23, 59))
        calculate_price.append(T_Day[0] - T_Day[-1])
        calculate_bz.append('当日差')
        calculate_ac_tg.append('当日差')

        calculate_date.append(it)
        calculate_time.append(datetime.time(23, 59))
        calculate_price.append(total_price[0] - total_price[-1])
        calculate_bz.append('结转差')
        calculate_ac_tg.append('结转差')

    res_date += calculate_date
    res_time += calculate_time
    res_bz += calculate_bz
    res_price += calculate_price
    res_ac_tg += calculate_ac_tg

    data = {
        'date': res_date, #6
        'time': res_time,#6
        'bz': res_bz,#6
        'price': res_price,#6
        'ac_tg': res_ac_tg#6
    }


    # 创建示例数据
    # data = {
    #     'date': ['A', 'B', 'A', 'B', 'A', 'B'], #6
    #     'time': ['00:00:00', '00:00:00', '00:00:00', '00:00:00', '00:00:00', '00:00:00'],#6
    #     'bz': ['', '', '', '', '', ''],#6
    #     'price': [100, 150, 200, 250, 300, 350],#6
    #     'ac_tg': ['A->X', 'B->X', 'C->X', 'X->A', 'X->D', 'X->D']#6
    # }

    # 将数据转换为 DataFrame
    df = pd.DataFrame(data)

    # 生成透视表
    pivot_table = pd.pivot_table(df, values='price', index=['date', 'time', 'bz'], columns='ac_tg', aggfunc='sum')

    pivot_table.to_excel(t_o_p, sheet_name='Pivot_Data')
    print('export xlsx done')

def foo(data_root_path, output_path):
    CD, BD = load_data(data_root_path)

    export_data([CD, BD], output_path)



if __name__ == '__main__':
    root_path = 'data/wjj'
    pj_path = 'pj1'
    output_path = 'output'

    foo(os.path.join(root_path, pj_path), os.path.join(root_path, output_path))