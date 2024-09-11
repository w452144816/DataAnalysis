import requests, json
from bs4 import BeautifulSoup
import openpyxl
from thefuzz import fuzz
import time
# 基类，后续可以在此之上扩展
class AbstractWebPage:
    def __init__(self, cookie, use_cookie=True):
        with open('cookie.txt') as f:
            cookie_str = f.readline()
        if use_cookie:
            self.headers = {

                "Authority": "api.m.jd.com",
                "Accept": "application / json, text / javascript, * / *; q = 0.01",
                "Accept - Encoding": "gzip, deflate, br",
                "Accept - Language": "zh - CN, zh;q = 0.9",
                "Cookie": "shshshfpb=i0ZU6VlHi9tt1RukWDDyR0w; 3AB9D23F7A4B3C9B=GZSZ6SPDPJZS6ARBGAUDIS7NMVC2A24XK6SN4JCWH44HGMYJVGXZIEY2SHDTJKNBR32WP5NA7JKC4CLDZDF5AIRXNA; shshshfpa=cb3af5e3-c2cf-dae5-48e3-c2331a38092a-1653253655; shshshfpx=cb3af5e3-c2cf-dae5-48e3-c2331a38092a-1653253655; __jdc=122270672; __jdv=122270672|direct|-|none|-|1689305241830; __jdu=16893052418291576334291; areaId=25; ipLoc-djd=25-2258-2261-6568; token=7a3a5010c8ea7250057d9168270daacd,2,939221; __tk=be32047e11adf495830ad564f7c34cd6,2,939221; 3AB9D23F7A4B3CSS=jdd03GZSZ6SPDPJZS6ARBGAUDIS7NMVC2A24XK6SN4JCWH44HGMYJVGXZIEY2SHDTRiDY9CRQSU93J9SUTiPmFy3PTP7N8itsNd7DLuiPzfoEjAAACXCBKUWUQMP7FMX; _gia_d=1; jsavif=1; __jda=122270672.16893052418291576334291.1689305242.1690550636.1690599310.7; __jdb=122270672.1.16893052418291576334291|7.1690599310",
                "Origin": "https://item.jd.com",
                "Referer": "https://item.jd.com/",
                "Sec-Ch-Ua": "\"Not.A/Brand\";v=\"8\", \"Chromium\";v=\"114\", \"Google Chrome\";v=\"114\"",
                "Sec-Ch-Ua-Mobile": "?0",
                "Sec-Ch-Ua-Platform": "\"Windows\"",
                "Sec-Fetch-Dest": "empty",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Site": "same-site",
                "X-Rp-Client": "h5_1.0.0",
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/80.0.3987.149 Safari/537.36',
                'cookie': cookie_str}
        else:
            self.headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/80.0.3987.149 Safari/537.36'
            }
        self.sess = requests.session()
        self.sess.headers.update(self.headers)


# 目录类，用来表示搜索结果
class Content(AbstractWebPage):
    def __init__(self, cookie, keyword, end_page):
        super(Content, self).__init__(cookie)
        self.keyword = keyword
        start_url = 'https://search.jd.com/Search?keyword=' + keyword + '&enc=utf-8&wq=' + keyword
        self.url_list = [start_url + '&page=' + str(j) for j in range(1, end_page + 1)]
        self.end_page = end_page

    def print(self):
        print(self.url_list, sep='\n')

    def get_item_info(self):
        item_pages_list = []
        with open("good_info.txt", 'w', encoding='utf-8') as f:
            f.write("产品名称" + '\t' + '价格' + '\t' + '销量' + '\t' '店铺' + '\n')
            f.write("*" * 50 + '\n')
            for url in self.url_list:
                isto = True
                while isto:
                    res = self.sess.get(url)

                    res.encoding = 'utf-8'
                    res = res.text
                    # 定位搜索结果主体，并获取所有的商品的标签
                    soup = BeautifulSoup(res, 'html.parser').select('#J_goodsList > ul')
                    if len(soup) != 0:
                        isto = False
                    time.sleep(1)
                good_list = soup[0].select('[class=gl-i-wrap]')
                # 循环获取所有商品信息
                for temp in good_list:
                    # 获取名称信息
                    name_div = temp.select_one('[class="p-name p-name-type-2"]')
                    tname = name_div.text.strip()
                    good_info = tname + '\t'

                    ratio_it = fuzz.partial_ratio(tname, self.keyword)
                    if ratio_it < 80:
                        continue

                    td = []
                    for link in name_div.find_all('a'):
                        td.append(link.get('href'))
                    # 价格信息
                    price_div = temp.select_one('[class=p-price]')
                    good_info += price_div.text.strip() + '\t'

                    # 评价信息
                    comment_div = temp.select_one('[class=p-commit]').find('strong').find('a')
                    comment_url = comment_div.get('href')
                    good_id = comment_url.replace('//item.jd.com/', '').replace('.html#comment', '')
                    # 评价信息没有在主页面内，而是需要另外发送GET获取，服务器地址如下
                    # 这里面的uuid是唯一标识符，如果运行程序发现报错或者没有得到想要的结果
                    # commit_start_url = f'https://api.m.jd.com/?appid=item-v3&functionId' \
                    #                    '=pc_club_productCommentSummaries&client=pc&clientVersion=1.0.0&t' \
                    #                    f'=1711091114924&referenceIds={good_id}&categoryIds=9987%2C653%2C655' \
                    #                    '&loginType=3&bbtf=&shield=&uuid=181111935.1679801589641754328424.1679801589' \
                    #                    '.1711082862.1711087044.29'
                    commit_start_url = f'https://api.m.jd.com/?appid=item-v3&functionId' \
                                       '=pc_club_productCommentSummaries&client=pc&clientVersion=1.0.0&t' \
                                       f'=1711091114924&referenceIds={good_id}&categoryIds=9987%2C653%2C655'
                    # 发送请求，得到结果

                    try:
                        comment_res = self.sess.get(commit_start_url)
                        # 编码方式是GBK国标编码
                        comment_res.encoding = 'gbk'
                        comment_res_json = comment_res.json()

                        # 解析得到评论数量
                        good_info += comment_res_json['CommentsCount'][0]['CommentCountStr'] + '\t'
                    except:
                        good_info += 'error \t'
                    # 店铺信息
                    shop_div = temp.select_one('[class=p-shop]')
                    good_info += shop_div.get_text().strip() + '\t'

                    good_info += td[0]
                    item_pages_list.append(good_info)
                    f.write(good_info + '\n')
                    f.write("*" * 50 + '\n')
            f.close()

        return '\n'.join(item_pages_list)


if __name__ == "__main__":
    # cookie，用于验证登录状态，必须要有cookie，否则京东会提示网络繁忙请重试
    # 获取方法：使用浏览器登录过后按F12，点击弹出界面中最上方的network选项，name栏里面随便点开一个，拉到最下面就有cookie，复制到cookie.txt中
    # 注意，不要换行，前后不要有空格，只需要复制cookie的值，不需要复制“cookie：”这几个字符
    # 上面的看不懂的话，看这个：https://blog.csdn.net/qq_46047971/article/details/121694916
    # 然后就可以运行程序了
    # cookie_str = ''
    # with open('cookie.txt') as f:
    #     cookie_str = f.readline()
    filen = 'jddata1.xlsx'
    workbook = openpyxl.load_workbook(filen)
    sheet = workbook['Sheet2']

    # 最大行 sheet.max_row # 最大列 sheet.max_column
    backdata = None
    for i in range(sheet.max_row + 1):
        if i < 2:
            continue
        spname = sheet.cell(row=i, column=2)
        print(spname.value)  # 获取内容用value方法
        sn = spname.value
        if '】' in spname.value:
            sn = ''.join(spname.value.split('】')[1:])
        content_page = Content('', sn, 1)
        content_page.print()

        urls = content_page.get_item_info()
        print(urls)
        sheet.cell(row=i, column=35, value=urls)
    # 输入cookie，关键词，输入结束页数
    workbook.save(filen)

