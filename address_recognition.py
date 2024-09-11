import requests
import json
import openpyxl

API_KEY = "hPcMmDIscCyS85OaNS2PrHOu"
SECRET_KEY = "LYuOXbSB45Fnm57k5XlXDNSXe7r5U19X"


def address_rec(_addr):
    url = "https://aip.baidubce.com/rpc/2.0/nlp/v1/address?access_token=" + get_access_token()

    payload = json.dumps({
        "text": _addr
    })
    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    print(response.text)
    return json.loads(response.text)


def get_access_token():
    """
    使用 AK，SK 生成鉴权签名（Access Token）
    :return: access_token，或是None(如果错误)
    """
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {"grant_type": "client_credentials", "client_id": API_KEY, "client_secret": SECRET_KEY}
    return str(requests.post(url, params=params).json().get("access_token"))


if __name__ == '__main__':
    # address_rec("四川省成都市郫都区望丛东路19号  刘春梅  15196646567")
    workbook = openpyxl.load_workbook('addr5.xlsx', data_only=True)
    sheet = workbook['Sheet1']

    wb = openpyxl.Workbook()
    # 选择默认的工作表
    wsheet = wb.active
    # 给工作表重命名
    wsheet.title = 'data'

    # 最大行 sheet.max_row # 最大列 sheet.max_column
    backdata = None
    for i in range(sheet.max_row + 1):
        if i < 2:
            if i == 1:
                wsheet.append(
                    ['省', '市', '区', '详细地址', '姓名', '电话'])
            continue
        data = sheet.cell(row=i, column=2)
        data_b1 = sheet.cell(row=i, column=3)
        data_b2 = sheet.cell(row=i, column=4)
        print(data.value)  # 获取内容用value方法
        writeData = []
        if data.value:
            tres = address_rec(data.value + ' ' + str(data_b1.value) + ' ' + str(data_b2.value))
            ts = list(tres.values())

            tname = ts[6] if len(ts[6]) > 0 else ts[1][-3:0]

            writeData = [ts[5], ts[11], ts[10], ts[2] + ts[1], tname, ts[3]]
            wsheet.append(
                writeData)
            backdata = writeData
        else:
            wsheet.append(
                backdata)


    # 保存 Excel 文件
    wb.save('res.xlsx')

